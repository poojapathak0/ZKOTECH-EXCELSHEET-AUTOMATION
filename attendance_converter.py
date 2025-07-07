import pandas as pd
import os
from datetime import datetime

def process_attendance_file(input_file_path, output_file_path=None):
    """
    Process Excel attendance file and convert to clean report format
    """
    try:
        # Read the Excel file with proper engine detection
        file_extension = input_file_path.split('.')[-1].lower()
        
        if file_extension == 'xls':
            # For .xls files, use xlrd engine
            df = pd.read_excel(input_file_path, engine='xlrd')
        elif file_extension in ['xlsx', 'xlsm']:
            # For .xlsx/.xlsm files, use openpyxl engine
            df = pd.read_excel(input_file_path, engine='openpyxl')
        else:
            # Try with openpyxl first, then xlrd
            try:
                df = pd.read_excel(input_file_path, engine='openpyxl')
            except:
                df = pd.read_excel(input_file_path, engine='xlrd')
        
        # Display original data info
        print("Original file loaded successfully!")
        print(f"Total students: {len(df)}")
        print(f"Columns: {list(df.columns)}")
        
        # Identify date columns (assuming they contain dates)
        date_columns = []
        for col in df.columns:
            if col not in ['Roll', 'Name', 'Present', 'Absent']:
                # Try to parse as date
                try:
                    pd.to_datetime(col)
                    date_columns.append(col)
                except:
                    # If it's not a date, check if it contains date-like patterns
                    if any(char in str(col) for char in ['/', '-', '2024', '2025']):
                        date_columns.append(col)
        
        print(f"Date columns found: {date_columns}")
        
        # Create a clean report
        report_data = []
        
        for index, row in df.iterrows():
            student_data = {
                'Roll No': row.get('Roll', ''),
                'Student Name': row.get('Name', ''),
                'Total Present': row.get('Present', 0),
                'Total Absent': row.get('Absent', 0)
            }
            
            # Add attendance for each date
            for date_col in date_columns:
                attendance_value = row.get(date_col, '')
                # Convert to P/A format
                if str(attendance_value).upper() in ['P', 'PRESENT']:
                    student_data[date_col] = 'P'
                elif str(attendance_value).upper() in ['A', 'ABSENT']:
                    student_data[date_col] = 'A'
                elif str(attendance_value).upper() in ['I', 'INCOMPLETE']:
                    student_data[date_col] = 'I'
                else:
                    student_data[date_col] = str(attendance_value)
            
            report_data.append(student_data)
        
        # Create DataFrame for the report
        report_df = pd.DataFrame(report_data)
        
        # Generate output filename if not provided
        if output_file_path is None:
            input_name = os.path.splitext(os.path.basename(input_file_path))[0]
            output_file_path = f"{input_name}_attendance_report.xlsx"
        
        # Save the report with formatting
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            report_df.to_excel(writer, sheet_name='Attendance Report', index=False)
            
            # Get the worksheet to apply formatting
            worksheet = writer.sheets['Attendance Report']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Apply basic formatting
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            
            # Header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            for col_num, cell in enumerate(worksheet[1], 1):
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Color code attendance cells
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    if cell.value == 'P':
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(color="006100")
                    elif cell.value == 'A':
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(color="9C0006")
                    elif cell.value == 'I':
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        cell.font = Font(color="9C5700")
                    
                    # Center align attendance values
                    if cell.column > 4:  # Attendance columns
                        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        print(f"Attendance report saved as: {output_file_path}")
        return output_file_path
        
    except Exception as e:
        print(f"Error processing file: {str(e)}")
        return None

def main():
    """
    Main function to run the attendance converter
    """
    print("=== Excel Attendance Report Converter ===")
    print()
    
    # Get input file path
    input_file = input("Enter the path to your Excel file (or drag and drop): ").strip().strip('"')
    
    if not os.path.exists(input_file):
        print("File not found! Please check the path.")
        return
    
    # Process the file
    output_file = process_attendance_file(input_file)
    
    if output_file:
        print(f"\n✅ Success! Attendance report created: {output_file}")
        print("The report includes:")
        print("- Color-coded attendance (Green=Present, Red=Absent, Yellow=Incomplete)")
        print("- Formatted headers and columns")
        print("- Summary of present/absent counts")
    else:
        print("\n❌ Failed to create attendance report")

if __name__ == "__main__":
    main()
