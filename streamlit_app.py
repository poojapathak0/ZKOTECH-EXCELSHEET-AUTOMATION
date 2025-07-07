import streamlit as st
import pandas as pd
import io
from datetime import datetime
import base64

def process_attendance_data(df):
    """Process the uploaded attendance dataframe"""
    
    # Show a clean preview first
    st.info("📋 **Analyzing your data structure...**")
    
    with st.expander("🔍 Data Analysis Details", expanded=False):
        st.write("**Column names found:**", list(df.columns))
        st.write("**Sample data (first 5 rows):**")
        st.dataframe(df.head())
    
    # Try to identify the structure of the data
    name_col = None
    roll_col = None
    date_col = None
    status_col = None
    time_col = None
    
    # Find column names (case insensitive)
    for col in df.columns:
        col_lower = str(col).lower()
        if 'name' in col_lower and not name_col:
            name_col = col
        elif 'roll' in col_lower and not roll_col:
            roll_col = col
        elif ('date' in col_lower or 'time' in col_lower or 
              any(char in str(col) for char in ['/', '-', '2024', '2025'])) and not date_col:
            date_col = col
        elif ('status' in col_lower or 'present' in col_lower or 
              'absent' in col_lower or 'attendance' in col_lower) and not status_col:
            status_col = col
        elif 'time' in col_lower and 'date' not in col_lower and not time_col:
            time_col = col
    
    # Smart fallback detection
    if not name_col and len(df.columns) >= 2:
        # Look for text columns that could be names
        for col in df.columns:
            if df[col].dtype == 'object' and not name_col:
                name_col = col
                break
    
    if not date_col:
        # Look for date-like data in columns
        for col in df.columns:
            if col not in [name_col, roll_col]:
                try:
                    pd.to_datetime(df[col].iloc[0], errors='raise')
                    date_col = col
                    break
                except:
                    continue
    
    if not status_col:
        # Look for status-like data
        for col in df.columns:
            if col not in [name_col, roll_col, date_col, time_col]:
                sample_values = df[col].astype(str).str.upper().unique()[:10]  # Check first 10 unique values
                if any(val in ['P', 'A', 'PRESENT', 'ABSENT', '1', '0'] for val in sample_values):
                    status_col = col
                    break
    
    # Show detected columns
    with st.expander("🎯 **Column Detection Results**", expanded=True):
        col1, col2 = st.columns(2)
        with col1:
            st.write("✅ **Detected Columns:**")
            st.write(f"• **Name:** {name_col or '❌ Not found'}")
            st.write(f"• **Roll:** {roll_col or '❌ Optional'}")
            st.write(f"• **Date:** {date_col or '❌ Not found'}")
        with col2:
            st.write("📊 **Data Columns:**")
            st.write(f"• **Status:** {status_col or '❌ Not found'}")
            st.write(f"• **Time:** {time_col or '❌ Optional'}")
    
    # Manual column selection if auto-detection fails
    if not all([name_col, date_col, status_col]):
        st.warning("⚠️ **Could not auto-detect all required columns. Please help us identify them:**")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            name_col = st.selectbox("📝 Select Name Column:", df.columns, 
                                  index=list(df.columns).index(name_col) if name_col else 0)
        with col2:
            date_col = st.selectbox("📅 Select Date Column:", df.columns,
                                  index=list(df.columns).index(date_col) if date_col else 1)
        with col3:
            status_col = st.selectbox("✅ Select Status Column:", df.columns,
                                    index=list(df.columns).index(status_col) if status_col else 2)
    
    try:
        # Create pivot table: Names as rows, Dates as columns, Status as values
        st.info("🔄 **Processing your data into attendance report format...**")
        
        df_clean = df.copy()
        
        # Convert date column to proper date format
        try:
            df_clean[date_col] = pd.to_datetime(df_clean[date_col]).dt.strftime('%m/%d/%Y')
        except:
            st.warning("⚠️ Could not parse dates, keeping original format")
        
        # Standardize status values - handle numeric values as well
        df_clean[status_col] = df_clean[status_col].astype(str).str.upper()
        
        # Check if values are numeric (likely student IDs) - convert to P if present, A if missing
        sample_values = df_clean[status_col].dropna().head(10).tolist()
        is_numeric = all(val.replace('.', '').replace('-', '').isdigit() or val == 'NAN' for val in sample_values)
        
        st.info(f"🔍 **Data Type Detection:** {'Numeric values detected (treating numbers as Present)' if is_numeric else 'Text values detected'}")
        
        if is_numeric:
            # If numeric values, treat any number as Present, NaN/empty as Absent
            df_clean[status_col] = df_clean[status_col].apply(lambda x: 'P' if x not in ['NAN', '', '-', 'NONE', 'nan'] and str(x).replace('.', '').replace('-', '').isdigit() else 'A')
        else:
            # Standard text-based conversion
            df_clean[status_col] = df_clean[status_col].replace({
                'PRESENT': 'P',
                'ABSENT': 'A',
                'INCOMPLETE': 'I',
                'TRUE': 'P',
                'FALSE': 'A',
                '1': 'P',
                '0': 'A'
            })
        
        # Show conversion sample
        with st.expander("🔄 **Status Conversion Preview**", expanded=False):
            conversion_sample = pd.DataFrame({
                'Original': df[status_col].head(10),
                'Converted': df_clean[status_col].head(10)
            })
            st.dataframe(conversion_sample)
        
        # Use just the status column (P/A only)
        value_col = status_col
        
        # Create pivot table
        index_cols = [roll_col, name_col] if roll_col and roll_col in df_clean.columns else [name_col]
        
        pivot_df = df_clean.pivot_table(
            index=index_cols,
            columns=date_col,
            values=value_col,
            aggfunc='first'  # Take first value if duplicates
        ).reset_index()
        
        # Flatten column names
        pivot_df.columns.name = None
        
        # Rename the first columns
        if roll_col and roll_col in pivot_df.columns:
            pivot_df = pivot_df.rename(columns={roll_col: 'Roll No', name_col: 'Student Name'})
        else:
            pivot_df = pivot_df.rename(columns={name_col: 'Student Name'})
            # Add a roll number column
            pivot_df.insert(0, 'Roll No', range(1, len(pivot_df) + 1))
        
        # Calculate totals
        date_columns = [col for col in pivot_df.columns if col not in ['Roll No', 'Student Name']]
        
        # Reorder columns first
        base_cols = ['Roll No', 'Student Name', 'Total Present', 'Total Absent']
        pivot_df = pivot_df[base_cols[:-2] + date_columns]  # Don't include Total columns yet
        
        # Fill NaN values with '-'
        pivot_df = pivot_df.fillna('-')
        
        # Now calculate totals after filling NaN values
        pivot_df['Total Present'] = 0
        pivot_df['Total Absent'] = 0
        
        for col in date_columns:
            # Count P as present
            pivot_df['Total Present'] += (pivot_df[col] == 'P').astype(int)
            # Count A as absent  
            pivot_df['Total Absent'] += (pivot_df[col] == 'A').astype(int)
            # Also count '-' (missing data) as absent
            pivot_df['Total Absent'] += (pivot_df[col] == '-').astype(int)
        
        # Reorder columns with totals
        base_cols = ['Roll No', 'Student Name', 'Total Present', 'Total Absent']
        pivot_df = pivot_df[base_cols + date_columns]
        
        st.success("✅ **Data processed successfully!**")
        return pivot_df
        
    except Exception as e:
        st.error(f"❌ **Error processing data:** {str(e)}")
        
        with st.expander("🆘 **Need Help?**", expanded=True):
            st.markdown("""
            **Please check your data format:**
            
            ✅ **Expected format for attendance data:**
            - Each row = one attendance record
            - Columns should include: Student Name, Date, Status (P/A)
            - Names can repeat for different dates
            
            📝 **Example:**
            ```
            Name        | Date       | Status
            John Doe    | 2025-01-01 | P
            John Doe    | 2025-01-02 | A
            Jane Smith  | 2025-01-01 | P
            ```
            """)
        
        return df

def style_dataframe(df):
    """Apply styling to the dataframe for better visualization"""
    
    def color_attendance(val):
        if val == 'P':
            return 'background-color: #d4edda; color: #155724; font-weight: bold; text-align: center; border: 1px solid #c3e6cb'
        elif val == 'A':
            return 'background-color: #f8d7da; color: #721c24; font-weight: bold; text-align: center; border: 1px solid #f5c6cb'
        elif val == 'I':
            return 'background-color: #fff3cd; color: #856404; font-weight: bold; text-align: center; border: 1px solid #ffeaa7'
        elif val == '-':
            return 'background-color: #e2e3e5; color: #383d41; font-weight: bold; text-align: center; border: 1px solid #d1d3d4'
        return ''
    
    # Apply styling to attendance columns (skip first 4 columns: Roll No, Name, Total Present, Total Absent)
    styled_df = df.style.applymap(color_attendance)
    return styled_df

def create_excel_download(df, filename):
    """Create Excel file for download"""
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Attendance Report', index=False)
        
        # Get worksheet for formatting
        worksheet = writer.sheets['Attendance Report']
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Apply formatting
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Color code attendance with softer, eye-friendly colors
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                if cell.value == 'P':
                    cell.fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
                    cell.font = Font(color="155724", bold=True)
                elif cell.value == 'A':
                    cell.fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
                    cell.font = Font(color="721c24", bold=True)
                elif cell.value == 'I':
                    cell.fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
                    cell.font = Font(color="856404", bold=True)
                elif cell.value == '-':
                    cell.fill = PatternFill(start_color="e2e3e5", end_color="e2e3e5", fill_type="solid")
                    cell.font = Font(color="383d41", bold=True)
                
                if cell.column > 4:  # Attendance columns
                    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    output.seek(0)
    return output

# Streamlit App
def main():
    st.set_page_config(
        page_title="Attendance Report Generator",
        page_icon="📊",
        layout="wide",
        initial_sidebar_state="collapsed"
    )
    
    # Custom CSS for better UI
    st.markdown("""
    <style>
    .main-header {
        text-align: center;
        padding: 2rem 0;
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
        color: white;
        border-radius: 10px;
        margin-bottom: 2rem;
    }
    .upload-section {
        background: #f8f9fa;
        padding: 2rem;
        border-radius: 10px;
        border: 2px dashed #dee2e6;
        text-align: center;
        margin: 2rem 0;
    }
    .metric-card {
        background: white;
        padding: 1rem;
        border-radius: 8px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        text-align: center;
    }
    .success-msg {
        background: #d4edda;
        color: #155724;
        padding: 1rem;
        border-radius: 8px;
        border: 1px solid #c3e6cb;
        margin: 1rem 0;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Main header
    st.markdown("""
    <div class="main-header">
        <h1>📊 Excel Attendance Report Generator</h1>
        <p>Transform your attendance data into professional reports with color-coded P/A indicators</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Instructions
    with st.container():
        st.markdown("### 🚀 **How to use:**")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown("""
            **1️⃣ Upload File**
            - Drag & drop your Excel file
            - Supports .xlsx and .xls formats
            - File size limit: 200MB
            """)
        
        with col2:
            st.markdown("""
            **2️⃣ Preview & Process**
            - Auto-detects your data structure
            - Converts to attendance report format
            - Shows color-coded P/A status
            """)
        
        with col3:
            st.markdown("""
            **3️⃣ Download Report**
            - Get formatted Excel report
            - Includes statistics & summaries
            - Professional styling applied
            """)
    
    st.markdown("---")
    
    # File upload section
    st.markdown("### 📁 **Upload your attendance data**")
    
    uploaded_file = st.file_uploader(
        "Choose your Excel file",
        type=['xlsx', 'xls'],
        help="Upload your attendance Excel file (max 200MB)",
        label_visibility="collapsed"
    )
    
    if uploaded_file is not None:
        try:
            # Read the uploaded file with proper engine detection
            with st.spinner("📖 Reading your Excel file..."):
                file_extension = uploaded_file.name.split('.')[-1].lower()
                
                if file_extension == 'xls':
                    df = pd.read_excel(uploaded_file, engine='xlrd')
                elif file_extension in ['xlsx', 'xlsm']:
                    df = pd.read_excel(uploaded_file, engine='openpyxl')
                else:
                    try:
                        df = pd.read_excel(uploaded_file, engine='openpyxl')
                    except:
                        df = pd.read_excel(uploaded_file, engine='xlrd')
            
            # Success message
            st.markdown(f"""
            <div class="success-msg">
                ✅ <strong>File uploaded successfully!</strong><br>
                📊 Found {len(df)} records in your file
            </div>
            """, unsafe_allow_html=True)
            
            # Show original data preview
            with st.expander("📋 **Original Data Preview**", expanded=False):
                st.dataframe(df.head(10), use_container_width=True)
            
            # Process the data
            with st.spinner("🔄 Processing your attendance data..."):
                report_df = process_attendance_data(df)
            
            # Only show results if processing was successful
            if report_df is not None and len(report_df) > 0:
                # Display statistics
                st.markdown("### 📊 **Report Statistics**")
                
                col1, col2, col3, col4 = st.columns(4)
                
                with col1:
                    st.metric(
                        label="👥 Total Students",
                        value=len(report_df)
                    )
                
                with col2:
                    if 'Total Present' in report_df.columns:
                        total_present = report_df['Total Present'].sum()
                        st.metric(
                            label="✅ Total Present",
                            value=int(total_present)
                        )
                    else:
                        st.metric(label="✅ Total Present", value="N/A")
                
                with col3:
                    if 'Total Absent' in report_df.columns:
                        total_absent = report_df['Total Absent'].sum()
                        st.metric(
                            label="❌ Total Absent (incl. No Data)",
                            value=int(total_absent)
                        )
                    else:
                        st.metric(label="❌ Total Absent (incl. No Data)", value="N/A")
                
                with col4:
                    date_columns = [col for col in report_df.columns if col not in ['Roll No', 'Student Name', 'Total Present', 'Total Absent']]
                    st.metric(
                        label="📅 Date Columns",
                        value=len(date_columns)
                    )
                
                st.markdown("---")
                
                # Display processed report
                st.markdown("### 📋 **Attendance Report Preview**")
                st.markdown("**Legend:** 🟢 **P** = Present | 🔴 **A** = Absent | 🟡 **I** = Incomplete | ⚫ **-** = No Data")
                st.markdown("**Note:** Both 'A' (Absent) and '-' (No Data) are counted as absent in totals.")
                
                # Add editing option
                edit_mode = st.toggle("✏️ **Enable Editing Mode**", value=False, help="Turn on to edit attendance data manually")
                
                if edit_mode:
                    st.info("📝 **Editing Mode Active:** You can now edit the attendance data directly. Changes will be reflected in the download.")
                    
                    # Create editable dataframe
                    # Configure columns for better editing experience
                    date_columns = [col for col in report_df.columns if col not in ['Roll No', 'Student Name', 'Total Present', 'Total Absent']]
                    column_config = {
                        "Roll No": st.column_config.NumberColumn(
                            "Roll No",
                            help="Student roll number",
                            disabled=True,
                            width="small"
                        ),
                        "Student Name": st.column_config.TextColumn(
                            "Student Name",
                            help="Student name",
                            disabled=True,
                            width="medium"
                        ),
                        "Total Present": st.column_config.NumberColumn(
                            "Total Present",
                            help="Total present days (auto-calculated)",
                            disabled=True,
                            width="small"
                        ),
                        "Total Absent": st.column_config.NumberColumn(
                            "Total Absent (incl. No Data)", 
                            help="Total absent days including 'No Data' entries (auto-calculated)",
                            disabled=True,
                            width="small"
                        )
                    }
                    
                    # Add selectbox configuration for date columns
                    for col in date_columns:
                        column_config[col] = st.column_config.SelectboxColumn(
                            col,
                            options=['P', 'A', 'I', '-'],
                            help="P=Present, A=Absent, I=Incomplete, -=No Data",
                            width="small"
                        )
                    
                    edited_df = st.data_editor(
                        report_df,
                        use_container_width=True,
                        height=400,
                        column_config=column_config,
                        key="attendance_editor"
                    )
                    
                    # Recalculate totals when data is edited
                    if not edited_df.equals(report_df):
                        date_columns = [col for col in edited_df.columns if col not in ['Roll No', 'Student Name', 'Total Present', 'Total Absent']]
                        
                        # Recalculate individual student totals (row by row)
                        for index, row in edited_df.iterrows():
                            present_count = 0
                            absent_count = 0
                            
                            # Count P, A, and - for each student
                            for col in date_columns:
                                val = row[col]
                                if val == 'P':
                                    present_count += 1
                                elif val == 'A' or val == '-':
                                    absent_count += 1
                            
                            # Update the individual student's totals
                            edited_df.at[index, 'Total Present'] = present_count
                            edited_df.at[index, 'Total Absent'] = absent_count
                        
                        # Update the report_df for download
                        report_df = edited_df.copy()
                        st.success("✅ **Data updated!** Individual and overall totals have been recalculated.")
                        
                        # Show updated statistics (overall totals)
                        col1, col2 = st.columns(2)
                        with col1:
                            st.metric("Updated Total Present", int(edited_df['Total Present'].sum()))
                        with col2:
                            st.metric("Updated Total Absent (incl. No Data)", int(edited_df['Total Absent'].sum()))
                        
                        # Show color-coded preview of edited data
                        st.markdown("### 🎨 **Color-Coded Preview of Your Edits**")
                        styled_edited_df = style_dataframe(edited_df)
                        st.dataframe(styled_edited_df, use_container_width=True, height=300)
                    
                    st.markdown("**Editing Tips:**")
                    st.markdown("- **Click on attendance cells** to edit with dropdown menu: P, A, I, or -")
                    st.markdown("- **P** = Present, **A** = Absent, **I** = Incomplete, **-** = No Data")
                    st.markdown("- **Individual student totals** update automatically for each row")
                    st.markdown("- **Overall totals** update automatically when you make changes")
                    st.markdown("- **Color preview** shows below when you edit data")
                    st.markdown("- **Roll No and Student Name** cannot be edited")
                else:
                    # Style and display the dataframe (read-only)
                    styled_df = style_dataframe(report_df)
                    st.dataframe(styled_df, use_container_width=True, height=400)
                
                # Download section
                st.markdown("### 💾 **Download Your Report**")
                
                col1, col2 = st.columns([2, 1])
                
                with col1:
                    # Create download file
                    excel_file = create_excel_download(report_df, "attendance_report.xlsx")
                    
                    st.download_button(
                        label="📥 Download Attendance Report (Excel)",
                        data=excel_file,
                        file_name=f"attendance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        help="Download the formatted attendance report as an Excel file",
                        use_container_width=True
                    )
                
                with col2:
                    st.info("📄 **Report Features:**\n- Color-coded attendance\n- Professional formatting\n- Auto-adjusted columns\n- Summary statistics")
                
                # Summary statistics
                with st.expander("📈 **Detailed Student Statistics**", expanded=False):
                    summary_data = []
                    for _, row in report_df.iterrows():
                        date_cols = [col for col in report_df.columns if col not in ['Roll No', 'Student Name', 'Total Present', 'Total Absent']]
                        present_count = 0
                        absent_count = 0
                        
                        for col in date_cols:
                            val = row[col]
                            if val == 'P':
                                present_count += 1
                            elif val == 'A' or val == '-':
                                absent_count += 1
                        
                        total_days = present_count + absent_count
                        attendance_pct = round((present_count / total_days * 100), 2) if total_days > 0 else 0
                        
                        summary_data.append({
                            'Roll No': row['Roll No'],
                            'Student Name': row['Student Name'],
                            'Present Days': present_count,
                            'Absent Days': absent_count,
                            'Total Days': total_days,
                            'Attendance %': attendance_pct
                        })
                    
                    summary_df = pd.DataFrame(summary_data)
                    st.dataframe(summary_df, use_container_width=True)
        
        except Exception as e:
            st.error(f"❌ **Error processing file:** {str(e)}")
            
            with st.expander("🔧 **Troubleshooting Help**", expanded=True):
                st.markdown("""
                **Common issues and solutions:**
                
                1. **File format issues:** Make sure your file is a valid Excel file (.xlsx or .xls)
                2. **Column names:** Ensure your file has columns for Name, Date, and Status
                3. **Data format:** Each row should represent one attendance entry
                4. **File size:** Keep file size under 200MB
                
                **Need help?** Check the original data preview above to verify your file structure.
                """)
    
    else:
        # Show example when no file is uploaded
        st.markdown("### 📝 **Expected Data Format**")
        
        example_data = pd.DataFrame({
            'Name': ['John Doe', 'John Doe', 'Jane Smith', 'Jane Smith'],
            'Date': ['2025-01-01', '2025-01-02', '2025-01-01', '2025-01-02'],
            'Status': ['Present', 'Absent', 'Present', 'Present'],
            'Time': ['09:30 AM', '-', '09:25 AM', '09:35 AM']
        })
        
        st.markdown("**Example of expected input format:**")
        st.dataframe(example_data, use_container_width=True)
        
        st.markdown("**This will be converted to:**")
        
        result_example = pd.DataFrame({
            'Roll No': [1, 2],
            'Student Name': ['John Doe', 'Jane Smith'],
            'Total Present': [1, 2],
            'Total Absent': [1, 0],
            '2025-01-01': ['P', 'P'],
            '2025-01-02': ['A', 'P']
        })
        
        st.dataframe(result_example, use_container_width=True)

if __name__ == "__main__":
    main()
